"""Serve a local browser interface for vehicle speed estimation."""

from __future__ import annotations

import json
import math
import csv
import io
import logging
import requests
import os
import threading
import time
import uuid
from collections import defaultdict, deque
from collections import Counter
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

from flask import Flask, Response, jsonify, render_template, request, send_from_directory, stream_with_context
from werkzeug.utils import secure_filename

BASE_DIR = Path(__file__).resolve().parent
UPLOAD_DIR = BASE_DIR / "data" / "uploads"
RESULT_DIR = BASE_DIR / "data" / "results"
LOG_DIR = BASE_DIR / "data" / "logs"
CAMERA_CONFIG_PATH = BASE_DIR / "data" / "camera-config.json"
ALLOWED_EXTENSIONS = {"mp4", "mov", "avi", "mkv", "m4v"}
DEFAULT_SOURCE = "1252,787;2298,803;5039,2159;-550,2159"
YOLO_MODEL_PATH = BASE_DIR / "truck-hauler-ft.pt"
BG_SUB_DEFAULTS = {
    "history": 500,
    "var_threshold": 50,
    "warmup_frames": 300,
    "sky_fraction": 0.15,
    "bottom_fraction": 0.85,
    "min_area": 300,
    "max_area": 8000,
    "min_aspect": 0.5,
    "max_aspect": 3.5,
    "min_solidity": 0.40,
    "merge_dist": 60,
}
try:
    LOG_TIMEZONE = ZoneInfo("Asia/Jakarta")
except ZoneInfoNotFoundError:
    # Windows may not ship the IANA timezone database; Jakarta is UTC+7 year-round.
    LOG_TIMEZONE = timezone(timedelta(hours=7))
TRACK_CONFIRMATION_FRAMES = 5
PRODUCTION_SHEET_ID = "1uzeFePgF0vGwEaEx61ZzCe3qlvew69o5hs0Q3oFdEzQ"
PRODUCTION_SHEET_NAME = "DATA"
MATERIAL_COLUMNS = ("RP", "ON", "FD", "BL", "TS", "MP", "MC", "BD", "CL", "N P")
MAX_MULTI_CAMERAS = 16

app = Flask(__name__)
logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger("pit_dispatch_monitoring")
app.config["MAX_CONTENT_LENGTH"] = 2 * 1024 * 1024 * 1024
jobs: dict[str, dict[str, Any]] = {}
jobs_lock = threading.Lock()
Point = tuple[float, float]


def allowed_video(filename: str) -> bool:
    """Return whether a filename has a supported video extension."""
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def parse_points(value: str, name: str, minimum: int, maximum: int | None = None) -> list[list[float]]:
    """Parse a semicolon-separated list of x,y coordinate pairs."""
    points = []
    for pair in value.split(";"):
        x_value, y_value = pair.strip().split(",", maxsplit=1)
        points.append([float(x_value), float(y_value)])
    if len(points) < minimum or (maximum is not None and len(points) > maximum):
        expected = f"{minimum}" if maximum == minimum else f"minimal {minimum}"
        raise ValueError(f"{name} harus berisi {expected} titik koordinat.")
    return points


def project_to_polyline(point: Point, polyline: list[Point]) -> tuple[float, float]:
    """Return distance along a polyline and the point's nearest offset in pixels."""
    import math

    best_distance = 0.0
    best_offset = float("inf")
    travelled = 0.0
    for start, end in zip(polyline, polyline[1:]):
        delta_x, delta_y = end[0] - start[0], end[1] - start[1]
        segment_length = math.hypot(delta_x, delta_y)
        if segment_length == 0:
            continue
        ratio = max(0.0, min(1.0, ((point[0] - start[0]) * delta_x + (point[1] - start[1]) * delta_y) / segment_length**2))
        projected_x, projected_y = start[0] + ratio * delta_x, start[1] + ratio * delta_y
        offset = math.hypot(point[0] - projected_x, point[1] - projected_y)
        if offset < best_offset:
            best_offset = offset
            best_distance = travelled + ratio * segment_length
        travelled += segment_length
    return best_distance, best_offset


def parse_gate_definitions(value: str) -> tuple[list[tuple[Point, Point]], list[float]]:
    """Parse ordered gate lines and distances between adjacent gates from JSON."""
    payload = json.loads(value)
    raw_gates = payload["gates"]
    raw_distances = payload["distances"]
    if len(raw_gates) < 2 or len(raw_distances) != len(raw_gates) - 1:
        raise ValueError("Setidaknya dua gate dan jarak antar-gate diperlukan.")
    gates: list[tuple[Point, Point]] = []
    for raw_gate in raw_gates:
        if len(raw_gate) != 2 or any(len(point) != 2 for point in raw_gate):
            raise ValueError("Setiap gate harus memiliki tepat dua titik.")
        start = (float(raw_gate[0][0]), float(raw_gate[0][1]))
        end = (float(raw_gate[1][0]), float(raw_gate[1][1]))
        gates.append((start, end))
    distances = [float(distance) for distance in raw_distances]
    if any(distance <= 0 for distance in distances):
        raise ValueError("Semua jarak antar-gate harus lebih besar dari nol.")
    return gates, distances


def cross_product(first: Point, second: Point, third: Point) -> float:
    """Return the signed turn formed by three two-dimensional points."""
    return (
        (second[0] - first[0]) * (third[1] - first[1])
        - (second[1] - first[1]) * (third[0] - first[0])
    )


def point_on_segment(point: Point, start: Point, end: Point) -> bool:
    """Return whether a collinear point falls within a finite line segment."""
    return (
        min(start[0], end[0]) <= point[0] <= max(start[0], end[0])
        and min(start[1], end[1]) <= point[1] <= max(start[1], end[1])
    )


def segments_intersect(first_start: Point, first_end: Point, second_start: Point, second_end: Point) -> bool:
    """Return whether a tracked movement segment crosses a finite gate segment."""
    first_start_turn = cross_product(first_start, first_end, second_start)
    first_end_turn = cross_product(first_start, first_end, second_end)
    second_start_turn = cross_product(second_start, second_end, first_start)
    second_end_turn = cross_product(second_start, second_end, first_end)
    crosses = (first_start_turn > 0) != (first_end_turn > 0) and (second_start_turn > 0) != (second_end_turn > 0)
    if crosses:
        return True
    return (
        first_start_turn == 0 and point_on_segment(second_start, first_start, first_end)
    ) or (
        first_end_turn == 0 and point_on_segment(second_end, first_start, first_end)
    ) or (
        second_start_turn == 0 and point_on_segment(first_start, second_start, second_end)
    ) or (
        second_end_turn == 0 and point_on_segment(first_end, second_start, second_end)
    )


def update_job(job_id: str, **values: Any) -> None:
    """Atomically merge updated processing state into a job record."""
    with jobs_lock:
        jobs[job_id].update(values)


def stop_requested(job_id: str) -> bool:
    """Return whether the operator requested an RTSP job to stop."""
    with jobs_lock:
        return bool(jobs.get(job_id, {}).get("stop_requested"))


def consume_log_snapshot_request(job_id: str) -> bool:
    with jobs_lock:
        job = jobs.get(job_id)
        if not job or not job.get("snapshot_requested"):
            return False
        job["snapshot_requested"] = False
        return True


def update_connection(job_id: str, status: str, message: str) -> None:
    """Update the connection state shown for an RTSP job."""
    update_job(job_id, connection_status=status, connection_message=message)
    logger.info("RTSP %s [%s] %s", job_id, status, message)


def mux_h264_frame(output_container: Any, output_stream: Any, frame: Any) -> None:
    """Encode one BGR frame into a browser-compatible H.264 video stream."""
    import av

    video_frame = av.VideoFrame.from_ndarray(frame, format="bgr24")
    for packet in output_stream.encode(video_frame):
        output_container.mux(packet)


def finish_h264_encoding(output_container: Any, output_stream: Any) -> None:
    """Flush delayed H.264 frames and close the output container safely."""
    for packet in output_stream.encode():
        output_container.mux(packet)
    output_container.close()


def format_log_sheet(worksheet: Any) -> None:
    """Apply consistent header, width, and frozen-row formatting to a log sheet."""
    from openpyxl.styles import Font, PatternFill

    header_fill = PatternFill("solid", fgColor="006F62")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for column_cells in worksheet.columns:
        column_letter = column_cells[0].column_letter
        longest_value = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[column_letter].width = min(max(longest_value + 2, 12), 28)


def write_detection_workbook(
    target_path: Path,
    source_video_name: str,
    mode: str,
    video_fps: float,
    video_frame_count: int,
    detection_logs: list[dict[str, Any]],
    speed_logs: list[dict[str, Any]],
) -> None:
    """Write detection and speed measurements to a formatted Excel workbook."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    workbook = Workbook()
    summary = workbook.active
    summary.title = "Ringkasan"
    summary.append(["Ringkasan Analisis Kecepatan", ""])
    summary.merge_cells("A1:B1")
    summary["A1"].font = Font(color="FFFFFF", bold=True, size=14)
    summary["A1"].fill = PatternFill("solid", fgColor="006F62")
    summary.append(["Video sumber", source_video_name])
    summary.append(["Metode", mode.title()])
    summary.append(["FPS", video_fps])
    summary.append(["Frame diproses", video_frame_count])
    summary.append(["Jumlah deteksi", len(detection_logs)])
    summary.append(["Pengukuran kecepatan", len(speed_logs)])
    summary.column_dimensions["A"].width = 28
    summary.column_dimensions["B"].width = 32

    vehicle_sheet = workbook.create_sheet("Ringkasan Kendaraan")
    vehicle_sheet.append(
        [
            "Tracker ID",
            "Frame pertama",
            "Frame terakhir",
            "Jumlah sampel",
            "Confidence rata-rata",
            "Kecepatan maksimum (km/h)",
        ]
    )
    vehicle_rows: dict[int, dict[str, Any]] = {}
    for entry in detection_logs:
        tracker_id = int(entry["tracker_id"])
        row = vehicle_rows.setdefault(
            tracker_id,
            {
                "first_frame": entry["frame"],
                "last_frame": entry["frame"],
                "samples": 0,
                "confidence_total": 0.0,
                "max_speed": 0,
            },
        )
        row["last_frame"] = entry["frame"]
        row["samples"] += 1
        row["confidence_total"] += entry["confidence"]
        if entry["speed_kmh"] is not None:
            row["max_speed"] = max(row["max_speed"], int(entry["speed_kmh"]))
    for tracker_id, row in sorted(vehicle_rows.items()):
        vehicle_sheet.append(
            [
                tracker_id,
                row["first_frame"],
                row["last_frame"],
                row["samples"],
                row["confidence_total"] / row["samples"],
                row["max_speed"],
            ]
        )
    format_log_sheet(vehicle_sheet)
    for cell in vehicle_sheet["E"][1:]:
        cell.number_format = "0.00"

    detections_sheet = workbook.create_sheet("Deteksi")
    detections_sheet.append(
        [
            "Frame",
            "Waktu video (detik)",
            "Tracker ID",
            "Kelas ID",
            "Confidence",
            "X1",
            "Y1",
            "X2",
            "Y2",
            "Mode",
            "Kecepatan terakhir (km/h)",
        ]
    )
    for entry in detection_logs:
        detections_sheet.append(
            [
                entry["frame"],
                entry["time_seconds"],
                entry["tracker_id"],
                entry["class_id"],
                entry["confidence"],
                entry["x1"],
                entry["y1"],
                entry["x2"],
                entry["y2"],
                entry["mode"],
                entry["speed_kmh"],
            ]
        )
    format_log_sheet(detections_sheet)
    for column in ("B", "E"):
        for cell in detections_sheet[column][1:]:
            cell.number_format = "0.00"

    speed_sheet = workbook.create_sheet("Kecepatan")
    speed_sheet.append(
        [
            "Frame",
            "Waktu video (detik)",
            "Tracker ID",
            "Segmen",
            "Jarak (meter)",
            "Waktu tempuh (detik)",
            "Kecepatan (km/h)",
        ]
    )
    for entry in speed_logs:
        speed_sheet.append(
            [
                entry["frame"],
                entry["time_seconds"],
                entry["tracker_id"],
                entry["segment"],
                entry["distance_meters"],
                entry["elapsed_seconds"],
                entry["speed_kmh"],
            ]
        )
    format_log_sheet(speed_sheet)
    for column in ("B", "E", "F", "G"):
        for cell in speed_sheet[column][1:]:
            cell.number_format = "0.00"
    workbook.save(target_path)


def record_detections(
    detection_logs: list[dict[str, Any]],
    detections: Any,
    frame_index: int,
    video_fps: float,
    mode: str,
    speeds: dict[int, int],
    confirmed_tracker_ids: set[int] | None = None,
) -> None:
    """Append only stable tracked detections and their latest speeds to the log."""
    for index, tracker_id in enumerate(detections.tracker_id):
        tracker_key = int(tracker_id)
        if confirmed_tracker_ids is not None and tracker_key not in confirmed_tracker_ids:
            continue
        x1, y1, x2, y2 = detections.xyxy[index]
        detection_logs.append(
            {
                "frame": frame_index,
                "time_seconds": frame_index / video_fps,
                "tracker_id": tracker_key,
                "class_id": int(detections.class_id[index]),
                "confidence": float(detections.confidence[index]),
                "x1": round(float(x1), 2),
                "y1": round(float(y1), 2),
                "x2": round(float(x2), 2),
                "y2": round(float(y2), 2),
                "mode": mode,
                "speed_kmh": speeds.get(tracker_key),
            }
        )


def parse_bg_sub_config(form_data: dict[str, Any]) -> dict[str, float | int]:
    """Parse and validate background subtraction settings from form data."""
    config: dict[str, float | int] = {}
    for key, default in BG_SUB_DEFAULTS.items():
        raw = form_data.get(key, "").strip()
        try:
            value = float(raw) if raw else float(default)
        except (TypeError, ValueError) as error:
            raise ValueError(f"Nilai {key} tidak valid.") from error
        if key in ("history", "warmup_frames"):
            value = int(value)
            if value < 0:
                raise ValueError(f"{key} harus lebih besar dari atau sama dengan nol.")
        elif key in ("min_area", "var_threshold") and value <= 0:
            raise ValueError(f"{key} harus lebih besar dari nol.")
        elif key in ("sky_fraction", "bottom_fraction", "min_solidity") and not 0 <= value <= 1:
            raise ValueError(f"{key} harus berada dalam rentang 0 sampai 1.")
        elif key == "min_aspect" and value <= 0:
            raise ValueError(f"{key} harus lebih besar dari nol.")
        elif key == "max_area" and value <= 0:
            raise ValueError(f"{key} harus lebih besar dari nol.")
        elif key == "merge_dist" and value < 0:
            raise ValueError(f"{key} harus lebih besar dari atau sama dengan nol.")
        config[key] = value
    if config["min_area"] >= config["max_area"]:
        raise ValueError("min_area harus lebih kecil dari max_area.")
    if config["min_aspect"] >= config["max_aspect"]:
        raise ValueError("min_aspect harus lebih kecil dari max_aspect.")
    if config["sky_fraction"] >= config["bottom_fraction"]:
        raise ValueError("sky_fraction harus lebih kecil dari bottom_fraction.")
    return config


def run_estimation(
    job_id: str,
    source_path: Path | str,
    target_path: Path,
    source_video_name: str,
    mode: str,
    corridor_polygon: list[list[float]] | None,
    route_points: list[list[float]] | None,
    route_length_meters: float | None,
    gates: list[tuple[Point, Point]] | None,
    gate_distances: list[float] | None,
    confidence_threshold: float,
    iou_threshold: float,
    detector: str = "yolo",
    bg_sub_config: dict[str, float | int] | None = None,
) -> None:
    """Run detection, tracking, and speed annotation for one uploaded video."""
    try:
        import cv2
        import numpy as np
        import supervision as sv
        import av

        is_rtsp = isinstance(source_path, str) and source_path.lower().startswith("rtsp://")
        capture = None
        if is_rtsp:
            capture = cv2.VideoCapture(source_path)
            if not capture.isOpened():
                raise ValueError("CCTV RTSP tidak dapat dibuka. Periksa URL, jaringan, dan kredensial kamera.")
            update_connection(job_id, "connected", "Koneksi RTSP tersambung.")
            fps = capture.get(cv2.CAP_PROP_FPS) or 25.0
            width = int(capture.get(cv2.CAP_PROP_FRAME_WIDTH))
            height = int(capture.get(cv2.CAP_PROP_FRAME_HEIGHT))
            if width <= 0 or height <= 0:
                raise ValueError("Resolusi stream RTSP tidak dapat dibaca.")
            from types import SimpleNamespace

            video_info = SimpleNamespace(
                fps=max(float(fps), 1.0),
                resolution_wh=(width, height),
                total_frames=0,
            )
        else:
            video_info = sv.VideoInfo.from_video_path(video_path=str(source_path))
        if mode == "polygon":
            if corridor_polygon is None or route_points is None or route_length_meters is None:
                raise ValueError("Koridor atau lintasan belum lengkap.")
            corridor = np.array(corridor_polygon, dtype=np.int32)
            route = [(float(x), float(y)) for x, y in route_points]
            route_pixel_total = sum(
                float(np.linalg.norm(np.subtract(end, start)))
                for start, end in zip(route, route[1:])
            )
            if route_pixel_total <= 0:
                raise ValueError("Lintasan harus memiliki panjang lebih dari nol.")
            meters_per_route_pixel = route_length_meters / route_pixel_total
        elif mode == "gate":
            if gates is None or gate_distances is None:
                raise ValueError("Kalibrasi gate belum lengkap.")
        else:
            raise ValueError("Metode pengukuran tidak dikenali.")

        model = None
        bg_detector = None
        if detector == "yolo":
            from ultralytics import YOLO

            if not YOLO_MODEL_PATH.is_file():
                raise FileNotFoundError(
                    f"Bobot model hauler tidak ditemukan: {YOLO_MODEL_PATH.name}"
                )
            model = YOLO(str(YOLO_MODEL_PATH))
        elif detector == "bg":
            from bg_subtraction import BgSubDetector

            bg_detector = BgSubDetector(**(bg_sub_config or {}))
        else:
            raise ValueError("Metode deteksi tidak dikenali.")

        tracker = sv.ByteTrack(
            frame_rate=video_info.fps,
            track_activation_threshold=confidence_threshold,
            lost_track_buffer=120 if bg_detector is not None else 30,
        )
        thickness = sv.calculate_optimal_line_thickness(video_info.resolution_wh)
        text_scale = sv.calculate_optimal_text_scale(video_info.resolution_wh)
        trace_annotator = sv.TraceAnnotator(
            thickness=thickness,
            trace_length=int(video_info.fps * 2),
            position=sv.Position.BOTTOM_CENTER,
        )
        box_annotator = sv.BoxAnnotator(thickness=thickness)
        label_annotator = sv.LabelAnnotator(
            text_scale=text_scale,
            text_thickness=thickness,
            text_position=sv.Position.BOTTOM_CENTER,
        )
        coordinates: defaultdict[int, deque[tuple[int, float]]] = defaultdict(
            lambda: deque(maxlen=max(1, int(video_info.fps)))
        )
        previous_positions: dict[int, Point] = {}
        smoothed_positions: dict[int, Point] = {}
        gate_entry_states: dict[int, tuple[int, int]] = {}
        completed_speeds: dict[int, int] = {}
        logged_speeds: dict[int, int] = {}
        detection_logs: list[dict[str, Any]] = []
        speed_logs: list[dict[str, Any]] = []
        track_hits: defaultdict[int, int] = defaultdict(int)
        frame_count = max(video_info.total_frames, 1)
        log_date: date = datetime.now(LOG_TIMEZONE).date()
        log_paths: list[Path] = []

        def daily_log_path(report_date: date) -> Path:
            if is_rtsp:
                return LOG_DIR / f"{job_id}-{report_date.isoformat()}-detection-log.xlsx"
            return LOG_DIR / f"{job_id}-detection-log.xlsx"

        def save_current_log(report_date: date) -> Path:
            path = daily_log_path(report_date)
            write_detection_workbook(
                path,
                source_video_name,
                mode,
                video_info.fps,
                video_info.total_frames or frame_index,
                detection_logs,
                speed_logs,
            )
            if path not in log_paths:
                log_paths.append(path)
            update_job(
                job_id,
                log_url=f"/logs/{path.name}",
                log_revision=datetime.now(timezone.utc).isoformat(),
            )
            return path

        output_container = None
        output_stream = None
        if not is_rtsp:
            output_container = av.open(str(target_path), mode="w", options={"movflags": "+faststart"})
            output_stream = output_container.add_stream("libx264", rate=round(video_info.fps))
            output_stream.width, output_stream.height = video_info.resolution_wh
            output_stream.pix_fmt = "yuv420p"
            output_stream.options = {"crf": "23", "preset": "veryfast"}
        frame_index = 0
        try:
            if is_rtsp:
                def rtsp_frames():
                    while not stop_requested(job_id):
                        ok, next_frame = capture.read()
                        if not ok:
                            update_connection(job_id, "reconnecting", "Stream terputus, mencoba reconnect...")
                            capture.release()
                            time.sleep(1)
                            if capture.open(source_path):
                                update_connection(job_id, "connected", "Koneksi RTSP tersambung kembali.")
                            else:
                                update_connection(job_id, "reconnecting", "Reconnect gagal, mencoba lagi...")
                            continue
                        yield next_frame

                frames = rtsp_frames()
            else:
                frames = sv.get_video_frames_generator(source_path=str(source_path))
            for frame_index, frame in enumerate(frames, start=1):
                if is_rtsp:
                    frame_date = datetime.now(LOG_TIMEZONE).date()
                    if frame_date != log_date:
                        save_current_log(log_date)
                        detection_logs.clear()
                        speed_logs.clear()
                        logged_speeds.clear()
                        log_date = frame_date
                    if consume_log_snapshot_request(job_id):
                        save_current_log(log_date)
                if bg_detector is not None:
                    detections = bg_detector.process(frame)
                    if detections is None:
                        if is_rtsp:
                            ok, encoded = cv2.imencode(".jpg", frame, [cv2.IMWRITE_JPEG_QUALITY, 85])
                            if ok:
                                update_job(job_id, latest_frame=encoded.tobytes())
                        else:
                            mux_h264_frame(output_container, output_stream, frame)
                            update_job(job_id, progress=round(frame_index / frame_count * 100))
                        continue
                else:
                    result = model(frame, conf=confidence_threshold, iou=iou_threshold, verbose=False)[0]
                    detections = sv.Detections.from_ultralytics(result)
                if mode == "polygon" and len(detections):
                    anchors = detections.get_anchors_coordinates(sv.Position.BOTTOM_CENTER)
                    inside = np.array(
                        [cv2.pointPolygonTest(corridor, (float(point[0]), float(point[1])), False) >= 0 for point in anchors]
                    )
                    detections = detections[inside]
                detections = tracker.update_with_detections(detections=detections)
                for tracker_id in detections.tracker_id:
                    track_hits[int(tracker_id)] += 1
                confirmed_tracker_ids = {
                    int(tracker_id)
                    for tracker_id in detections.tracker_id
                    if track_hits[int(tracker_id)] >= TRACK_CONFIRMATION_FRAMES
                }
                points = detections.get_anchors_coordinates(sv.Position.BOTTOM_CENTER)
                if bg_detector is not None:
                    smoothed = []
                    for tracker_id, point in zip(detections.tracker_id, points, strict=True):
                        tracker_key = int(tracker_id)
                        previous_smoothed = smoothed_positions.get(tracker_key)
                        if previous_smoothed is None:
                            current = (float(point[0]), float(point[1]))
                        else:
                            current = (
                                previous_smoothed[0] * 0.5 + float(point[0]) * 0.5,
                                previous_smoothed[1] * 0.5 + float(point[1]) * 0.5,
                            )
                        smoothed_positions[tracker_key] = current
                        smoothed.append(current)
                    points = np.asarray(smoothed, dtype=np.float64)

                labels = []
                if mode == "polygon":
                    for tracker_id, point in zip(detections.tracker_id, points, strict=True):
                        if int(tracker_id) not in confirmed_tracker_ids:
                            labels.append(f"#{tracker_id}")
                            continue
                        route_distance_px, route_offset_px = project_to_polyline(
                            (float(point[0]), float(point[1])), route
                        )
                        if route_offset_px > 150:
                            labels.append(f"#{tracker_id}")
                            continue
                        route_distance_m = route_distance_px * meters_per_route_pixel
                        coordinates[int(tracker_id)].append((frame_index, route_distance_m))
                        history = coordinates[int(tracker_id)]
                        if len(history) < video_info.fps / 2:
                            labels.append(f"#{tracker_id}")
                            continue
                        elapsed_time = (history[-1][0] - history[0][0]) / video_info.fps
                        distance = abs(history[-1][1] - history[0][1])
                        if elapsed_time <= 0:
                            labels.append(f"#{tracker_id}")
                            continue
                        speed = int(distance / elapsed_time * 3.6)
                        completed_speeds[int(tracker_id)] = speed
                        if logged_speeds.get(int(tracker_id)) != speed:
                            speed_logs.append(
                                {
                                    "frame": frame_index,
                                    "time_seconds": frame_index / video_info.fps,
                                    "tracker_id": int(tracker_id),
                                    "segment": "Area poligon",
                                    "distance_meters": distance,
                                    "elapsed_seconds": elapsed_time,
                                    "speed_kmh": speed,
                                }
                            )
                            logged_speeds[int(tracker_id)] = speed
                        labels.append(f"#{tracker_id} {speed} km/h")
                else:
                    for tracker_id, point in zip(detections.tracker_id, points, strict=True):
                        tracker_key = int(tracker_id)
                        if tracker_key not in confirmed_tracker_ids:
                            labels.append(f"#{tracker_id}")
                            continue
                        current_position = (float(point[0]), float(point[1]))
                        previous_position = previous_positions.get(tracker_key)
                        if previous_position is not None:
                            crossed_gate_indices = [
                                index
                                for index, gate in enumerate(gates)
                                if segments_intersect(
                                    previous_position, current_position, gate[0], gate[1]
                                )
                            ]
                            crossed_gate_index = crossed_gate_indices[0] if crossed_gate_indices else None
                            entry_state = gate_entry_states.get(tracker_key)
                            if entry_state is None and crossed_gate_index is not None:
                                gate_entry_states[tracker_key] = (crossed_gate_index, frame_index)
                            elif crossed_gate_index is not None and entry_state is not None:
                                previous_gate_index = entry_state[0]
                                gate_distance_index = abs(previous_gate_index - crossed_gate_index)
                                if gate_distance_index == 1:
                                    elapsed_time = (frame_index - entry_state[1]) / video_info.fps
                                    if elapsed_time > 0:
                                        distance_index = min(previous_gate_index, crossed_gate_index)
                                        completed_speeds[tracker_key] = int(
                                            gate_distances[distance_index] / elapsed_time * 3.6
                                        )
                                        speed_logs.append(
                                            {
                                                "frame": frame_index,
                                                "time_seconds": frame_index / video_info.fps,
                                                "tracker_id": tracker_key,
                                                "segment": f"Gate {chr(65 + previous_gate_index)} - Gate {chr(65 + crossed_gate_index)}",
                                                "distance_meters": gate_distances[distance_index],
                                                "elapsed_seconds": elapsed_time,
                                                "speed_kmh": completed_speeds[tracker_key],
                                            }
                                        )
                                    gate_entry_states[tracker_key] = (crossed_gate_index, frame_index)
                                elif gate_distance_index > 1:
                                    gate_entry_states[tracker_key] = (crossed_gate_index, frame_index)
                                else:
                                    gate_entry_states[tracker_key] = (crossed_gate_index, frame_index)
                        previous_positions[tracker_key] = current_position
                        speed = completed_speeds.get(tracker_key)
                        labels.append(f"#{tracker_id}" if speed is None else f"#{tracker_id} {speed} km/h")

                annotated_frame = trace_annotator.annotate(frame.copy(), detections)
                annotated_frame = box_annotator.annotate(annotated_frame, detections)
                annotated_frame = label_annotator.annotate(annotated_frame, detections, labels)
                if mode == "gate":
                    for index, gate in enumerate(gates):
                        gate_start = tuple(map(int, gate[0]))
                        gate_end = tuple(map(int, gate[1]))
                        cv2.line(annotated_frame, gate_start, gate_end, (80, 190, 80), thickness)
                        cv2.putText(
                            annotated_frame,
                            chr(65 + index),
                            gate_start,
                            cv2.FONT_HERSHEY_SIMPLEX,
                            text_scale,
                            (80, 190, 80),
                            thickness,
                        )
                elif mode == "polygon":
                    cv2.polylines(annotated_frame, [corridor], True, (80, 190, 80), thickness)
                    cv2.polylines(
                        annotated_frame,
                        [np.asarray(route, dtype=np.int32)],
                        False,
                        (80, 190, 80),
                        thickness,
                    )
                record_detections(
                    detection_logs,
                    detections,
                    frame_index,
                    video_info.fps,
                    mode,
                    completed_speeds,
                    confirmed_tracker_ids,
                )
                if is_rtsp:
                    ok, encoded = cv2.imencode(".jpg", annotated_frame, [cv2.IMWRITE_JPEG_QUALITY, 85])
                    if ok:
                        update_job(job_id, latest_frame=encoded.tobytes(), frames_processed=frame_index)
                else:
                    mux_h264_frame(output_container, output_stream, annotated_frame)
                    update_job(job_id, progress=round(frame_index / frame_count * 100))
        finally:
            if output_container is not None:
                finish_h264_encoding(output_container, output_stream)
            if capture is not None:
                capture.release()

        log_path = save_current_log(log_date)

        update_job(
            job_id,
            status="stopped" if is_rtsp and stop_requested(job_id) else "complete",
            progress=100,
            result_url=None if is_rtsp else f"/results/{target_path.name}",
            log_url=f"/logs/{log_path.name}",
            log_urls=[f"/logs/{path.name}" for path in log_paths],
            stream_url=f"/api/jobs/{job_id}/stream" if is_rtsp else None,
        )
    except Exception as error:
        if isinstance(source_path, str) and source_path.lower().startswith("rtsp://"):
            update_connection(job_id, "error", f"Koneksi RTSP gagal: {error}")
        update_job(job_id, status="error", error=str(error))


@app.get("/")
def index() -> str:
    """Render the video analysis workspace."""
    return render_template("index.html", default_source=DEFAULT_SOURCE)


@app.post("/api/jobs")
def create_job() -> tuple[Any, int] | Any:
    """Save an uploaded video and start its estimation job in a background thread."""
    source_type = request.form.get("source_type", "file")
    video = request.files.get("video")
    rtsp_url = request.form.get("rtsp_url", "").strip()
    if source_type == "rtsp":
        if not rtsp_url.lower().startswith("rtsp://"):
            return jsonify(error="URL CCTV harus diawali rtsp://."), 400
    elif video is None or not video.filename:
        return jsonify(error="Pilih file video terlebih dahulu."), 400
    elif not allowed_video(video.filename):
        return jsonify(error="Format video tidak didukung."), 400
    try:
        mode = request.form.get("mode", "polygon")
        detector = request.form.get("detector", "yolo")
        confidence_threshold = float(request.form["confidence_threshold"])
        iou_threshold = float(request.form["iou_threshold"])
        if detector not in {"yolo", "bg"}:
            raise ValueError("Metode deteksi tidak dikenali.")
        if not 0 < confidence_threshold <= 1 or not 0 < iou_threshold <= 1:
            raise ValueError("Nilai konfigurasi harus berada dalam rentang yang valid.")
        corridor_polygon = None
        route_points = None
        route_length_meters = None
        gates = None
        gate_distances = None
        bg_sub_config = None
        if detector == "bg":
            bg_sub_config = parse_bg_sub_config(request.form)
        if mode == "polygon":
            corridor_polygon = parse_points(request.form["corridor_polygon"], "Koridor jalan", 3)
            route_points = parse_points(request.form["route_points"], "Lintasan tengah", 2)
            route_length_meters = float(request.form["route_length_meters"])
            if route_length_meters <= 0:
                raise ValueError("Panjang lintasan harus lebih besar dari nol.")
        elif mode == "gate":
            gates, gate_distances = parse_gate_definitions(request.form["gate_definitions"])
        else:
            raise ValueError("Metode pengukuran tidak dikenali.")
    except (KeyError, TypeError, ValueError) as error:
        return jsonify(error=f"Konfigurasi tidak valid: {error}"), 400

    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    RESULT_DIR.mkdir(parents=True, exist_ok=True)
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    job_id = uuid.uuid4().hex
    filename = "CCTV RTSP" if source_type == "rtsp" else secure_filename(video.filename)
    source_path: Path | str = rtsp_url if source_type == "rtsp" else UPLOAD_DIR / f"{job_id}-{filename}"
    target_path = RESULT_DIR / f"{job_id}-estimated.mp4"
    if source_type != "rtsp":
        video.save(source_path)
    with jobs_lock:
        jobs[job_id] = {
            "status": "processing",
            "progress": 0,
            "error": None,
            "stream_url": f"/api/jobs/{job_id}/stream" if source_type == "rtsp" else None,
            "stop_requested": False,
            "connection_status": "connecting" if source_type == "rtsp" else None,
            "connection_message": "Menghubungkan ke RTSP..." if source_type == "rtsp" else None,
            "snapshot_requested": False,
            "log_url": None,
            "log_revision": None,
        }
    thread = threading.Thread(
        target=run_estimation,
        kwargs={
            "job_id": job_id,
            "source_path": source_path,
            "target_path": target_path,
            "source_video_name": filename,
            "mode": mode,
            "corridor_polygon": corridor_polygon,
            "route_points": route_points,
            "route_length_meters": route_length_meters,
            "gates": gates,
            "gate_distances": gate_distances,
            "confidence_threshold": confidence_threshold,
            "iou_threshold": iou_threshold,
            "detector": detector,
            "bg_sub_config": bg_sub_config,
        },
        daemon=True,
    )
    thread.start()
    return jsonify(job_id=job_id), 202


@app.post("/api/multi-rtsp-jobs")
def create_multi_rtsp_jobs() -> tuple[Any, int] | Any:
    """Start independently calibrated RTSP analysis jobs."""
    payload = request.get_json(silent=True) or {}
    cameras = payload.get("cameras")
    if not isinstance(cameras, list) or not 2 <= len(cameras) <= MAX_MULTI_CAMERAS:
        return jsonify(error=f"Masukkan dua sampai {MAX_MULTI_CAMERAS} kamera RTSP."), 400
    try:
        mode = payload["mode"]
        detector = payload["detector"]
        confidence_threshold = float(payload["confidence_threshold"])
        iou_threshold = float(payload["iou_threshold"])
        if mode not in {"polygon", "gate"} or detector not in {"yolo", "bg"}:
            raise ValueError("Mode atau metode deteksi tidak dikenali.")
        if not 0 < confidence_threshold <= 1 or not 0 < iou_threshold <= 1:
            raise ValueError("Nilai konfigurasi harus berada dalam rentang yang valid.")
        bg_sub_config = parse_bg_sub_config(payload) if detector == "bg" else None
        configured_cameras = []
        for index, camera in enumerate(cameras, start=1):
            url = str(camera.get("url", "")).strip()
            if not url.lower().startswith("rtsp://"):
                raise ValueError(f"URL Kamera {index} harus diawali rtsp://.")
            name = secure_filename(str(camera.get("name", "") or f"Kamera-{index}")) or f"Kamera-{index}"
            if mode == "polygon":
                corridor_polygon = camera.get("corridor_polygon")
                route_points = camera.get("route_points")
                route_length_meters = float(camera.get("route_length_meters"))
                if not isinstance(corridor_polygon, list) or len(corridor_polygon) < 3:
                    raise ValueError(f"Kalibrasi koridor Kamera {index} belum lengkap.")
                if (
                    not isinstance(route_points, list)
                    or len(route_points) < 2
                    or not math.isfinite(route_length_meters)
                    or route_length_meters <= 0
                ):
                    raise ValueError(f"Kalibrasi lintasan Kamera {index} belum lengkap.")
                gates = gate_distances = None
            else:
                gates, gate_distances = parse_gate_definitions(json.dumps(camera["gate_definitions"]))
                corridor_polygon = route_points = None
                route_length_meters = None
            configured_cameras.append(
                (name, url, corridor_polygon, route_points, route_length_meters, gates, gate_distances)
            )
    except (KeyError, TypeError, ValueError) as error:
        return jsonify(error=f"Konfigurasi multi-kamera tidak valid: {error}"), 400

    RESULT_DIR.mkdir(parents=True, exist_ok=True)
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    jobs_payload = []
    for name, url, corridor_polygon, route_points, route_length_meters, gates, gate_distances in configured_cameras:
        job_id = uuid.uuid4().hex
        target_path = RESULT_DIR / f"{job_id}-estimated.mp4"
        with jobs_lock:
            jobs[job_id] = {
                "status": "processing",
                "progress": 0,
                "error": None,
                "camera_name": name,
                "stream_url": f"/api/jobs/{job_id}/stream",
                "stop_requested": False,
                "connection_status": "connecting",
                "connection_message": "Menghubungkan ke RTSP...",
                "snapshot_requested": False,
                "log_url": None,
                "log_revision": None,
            }
        thread = threading.Thread(
            target=run_estimation,
            kwargs={
                "job_id": job_id,
                "source_path": url,
                "target_path": target_path,
                "source_video_name": name,
                "mode": mode,
                "corridor_polygon": corridor_polygon,
                "route_points": route_points,
                "route_length_meters": route_length_meters,
                "gates": gates,
                "gate_distances": gate_distances,
                "confidence_threshold": confidence_threshold,
                "iou_threshold": iou_threshold,
                "detector": detector,
                "bg_sub_config": bg_sub_config,
            },
            daemon=True,
        )
        thread.start()
        jobs_payload.append({"job_id": job_id, "camera_name": name})
    return jsonify(jobs=jobs_payload), 202


@app.post("/api/jobs/<job_id>/stop")
def stop_job(job_id: str) -> tuple[Any, int] | Any:
    """Request a running RTSP job to stop after its current frame."""
    with jobs_lock:
        job = jobs.get(job_id)
        if job is None:
            return jsonify(error="Pekerjaan tidak ditemukan."), 404
        if not job.get("stream_url"):
            return jsonify(error="Hanya stream RTSP yang dapat dihentikan."), 400
        job["stop_requested"] = True
    return jsonify(status="stopping"), 202


@app.post("/api/jobs/<job_id>/log-snapshot")
def request_log_snapshot(job_id: str) -> tuple[Any, int] | Any:
    """Queue an up-to-date Excel log snapshot for an RTSP job."""
    with jobs_lock:
        job = jobs.get(job_id)
        if job is None:
            return jsonify(error="Pekerjaan tidak ditemukan."), 404
        if not job.get("stream_url"):
            return jsonify(error="Snapshot log hanya tersedia untuk stream RTSP."), 400
        if job.get("status") not in {"processing", "stopped", "complete"}:
            return jsonify(error="Sesi RTSP belum siap membuat log."), 409
        if job.get("status") != "processing":
            return jsonify(status="ready", log_url=job.get("log_url")), 200
        job["snapshot_requested"] = True
        revision = job.get("log_revision")
    return jsonify(status="queued", previous_revision=revision), 202


@app.get("/api/camera-configurations")
def get_camera_configuration() -> Any:
    """Return the locally saved RTSP camera list without exposing it in source code."""
    if not CAMERA_CONFIG_PATH.is_file():
        return jsonify(cameras=[])
    try:
        payload = json.loads(CAMERA_CONFIG_PATH.read_text(encoding="utf-8"))
        cameras = payload.get("cameras", [])
        if not isinstance(cameras, list):
            raise ValueError
        return jsonify(cameras=cameras)
    except (OSError, ValueError, json.JSONDecodeError):
        return jsonify(error="Konfigurasi kamera lokal tidak dapat dibaca."), 500


@app.put("/api/camera-configurations")
def save_camera_configuration() -> tuple[Any, int] | Any:
    """Persist the current camera names and RTSP URLs on the local machine."""
    payload = request.get_json(silent=True) or {}
    raw_cameras = payload.get("cameras")
    if not isinstance(raw_cameras, list) or not 1 <= len(raw_cameras) <= MAX_MULTI_CAMERAS:
        return jsonify(error=f"Simpan satu sampai {MAX_MULTI_CAMERAS} kamera."), 400
    cameras = []
    for index, camera in enumerate(raw_cameras, start=1):
        name = str(camera.get("name", "") or f"Kamera {index}").strip()
        url = str(camera.get("url", "")).strip()
        if not url.lower().startswith("rtsp://"):
            return jsonify(error=f"URL Kamera {index} harus diawali rtsp://."), 400
        cameras.append({
            "name": name,
            "url": url,
            "corridorPoints": camera.get("corridorPoints", []),
            "routePoints": camera.get("routePoints", []),
            "calibrationPoints": camera.get("calibrationPoints", []),
            "polygonStep": camera.get("polygonStep", "corridor"),
            "gateCount": camera.get("gateCount", 2),
            "gateDistances": camera.get("gateDistances", [""]),
            "routeLength": camera.get("routeLength", ""),
        })
    try:
        CAMERA_CONFIG_PATH.parent.mkdir(parents=True, exist_ok=True)
        CAMERA_CONFIG_PATH.write_text(json.dumps({"cameras": cameras}, indent=2), encoding="utf-8")
    except OSError as error:
        return jsonify(error=f"Konfigurasi kamera tidak dapat disimpan: {error}"), 500
    return jsonify(cameras=cameras), 200


@app.get("/api/rtsp/preview")
def rtsp_preview() -> Any:
    """Return one JPEG frame so the operator can calibrate a CCTV stream."""
    import cv2

    url = request.args.get("url", "").strip()
    if not url.lower().startswith("rtsp://"):
        return jsonify(error="URL CCTV harus diawali rtsp://."), 400
    capture = cv2.VideoCapture(url)
    try:
        if not capture.isOpened():
            return jsonify(error="CCTV RTSP tidak dapat dibuka."), 400
        ok, frame = capture.read()
        if not ok:
            return jsonify(error="Frame CCTV tidak tersedia."), 400
        ok, encoded = cv2.imencode(".jpg", frame)
        if not ok:
            return jsonify(error="Frame CCTV gagal dikodekan."), 500
        return Response(encoded.tobytes(), mimetype="image/jpeg")
    finally:
        capture.release()


@app.get("/api/jobs/<job_id>/stream")
def job_stream(job_id: str) -> Response:
    """Stream the latest annotated RTSP frames as multipart JPEG."""
    def frames():
        import time

        last_frame = None
        while True:
            with jobs_lock:
                job = jobs.get(job_id)
                if job is None:
                    return
                current_frame = job.get("latest_frame")
                status = job.get("status")
            if current_frame is not None and current_frame != last_frame:
                last_frame = current_frame
                yield b"--frame\r\nContent-Type: image/jpeg\r\n\r\n" + current_frame + b"\r\n"
            if status in {"complete", "stopped", "error"} and current_frame == last_frame:
                return
            time.sleep(0.04)

    return Response(stream_with_context(frames()), mimetype="multipart/x-mixed-replace; boundary=frame")


@app.get("/api/jobs/<job_id>")
def get_job(job_id: str) -> tuple[Any, int] | Any:
    """Return the current state of a video processing job."""
    with jobs_lock:
        job = jobs.get(job_id)
    if job is None:
        return jsonify(error="Pekerjaan tidak ditemukan."), 404
    public_job = {key: value for key, value in job.items() if key != "latest_frame"}
    return jsonify(public_job)


@app.get("/results/<path:filename>")
def result_file(filename: str) -> Any:
    """Serve a completed analysis video from the local results directory."""
    return send_from_directory(RESULT_DIR, filename)


@app.get("/logs/<path:filename>")
def log_file(filename: str) -> Any:
    """Serve an Excel detection log from the local logs directory."""
    return send_from_directory(LOG_DIR, filename)


def report_file_path(filename: str) -> Path:
    """Return a validated workbook path from the local report directory."""
    path = LOG_DIR / secure_filename(filename)
    if path.suffix.lower() != ".xlsx" or not path.is_file():
        raise FileNotFoundError(filename)
    return path


def parse_sheet_number(value: str | None) -> float:
    """Parse the mixed Indonesian numeric formatting used by the production sheet."""
    raw = (value or "").strip().replace(" ", "").replace("-", "")
    if not raw:
        return 0.0
    if raw.endswith("x") or raw.endswith("X"):
        raw = raw[:-1].strip()
    if not raw:
        return 0.0
    if raw.endswith("%"):
        return parse_sheet_number(raw[:-1]) / 100
    if "," in raw:
        return float(raw.replace(".", "").replace(",", "."))
    if raw.count(".") > 1 or ("." in raw and len(raw.rsplit(".", 1)[1]) == 3):
        return float(raw.replace(".", ""))
    return float(raw)


def parse_sheet_date(value: str) -> datetime:
    """Parse the abbreviated Indonesian/English dates in the production sheet."""
    months = {"Jan": "Jan", "Feb": "Feb", "Mar": "Mar", "Apr": "Apr", "Mei": "May", "Jun": "Jun", "Jul": "Jul", "Agu": "Aug", "Sep": "Sep", "Okt": "Oct", "Nov": "Nov", "Des": "Dec"}
    day, month, year = value.strip().split("-", maxsplit=2)
    return datetime.strptime(f"{day}-{months[month]}-{year}", "%d-%b-%y")


def load_production_rows() -> list[dict[str, Any]]:
    """Load daily production rows from the public Google Sheet CSV export."""
    response = requests.get(
        f"https://docs.google.com/spreadsheets/d/{PRODUCTION_SHEET_ID}/gviz/tq",
        params={"tqx": "out:csv", "sheet": PRODUCTION_SHEET_NAME},
        timeout=20,
    )
    response.raise_for_status()
    rows = []
    for raw_row in csv.DictReader(io.StringIO(response.text)):
        row = {key.strip(): value for key, value in raw_row.items() if key is not None}
        try:
            row["parsed_date"] = parse_sheet_date(row["DATE"])
        except (KeyError, ValueError):
            continue
        rows.append(row)
    return rows


@app.get("/dashboard")
def dashboard() -> str:
    """Render the visual analytics dashboard for generated Excel reports."""
    return render_template("dashboard.html")


@app.get("/production-dashboard")
def production_dashboard() -> str:
    """Render the mining production dashboard backed by the public Sheet."""
    return render_template("production_dashboard.html")


@app.get("/api/production")
def production_data() -> tuple[Any, int] | Any:
    """Aggregate the public daily production sheet for the selected date range according to grouping rules."""
    try:
        rows = load_production_rows()
        start = request.args.get("start")
        end = request.args.get("end")
        if start:
            start_date = datetime.strptime(start, "%Y-%m-%d")
            rows = [row for row in rows if row["parsed_date"] >= start_date]
        if end:
            end_date = datetime.strptime(end, "%Y-%m-%d")
            rows = [row for row in rows if row["parsed_date"] <= end_date]
        if not rows:
            return jsonify(error="Tidak ada data pada rentang tanggal ini."), 404

        total = lambda column: round(sum(parse_sheet_number(row.get(column)) for row in rows), 2)
        avg = lambda column: round(sum(parse_sheet_number(row.get(column)) for row in rows) / len(rows), 2)

        # 1. Material Review (Hijau muda / Kode B - K)
        materials = [{"name": column, "value": total(column)} for column in MATERIAL_COLUMNS]

        # 2. Production Summary (OB, Coal, SR, Productivity)
        plan_ob = total("PLAN OB")
        actual_ob = total("ACT OB")
        plan_coal = total("PLAN COAL")
        actual_coal = total("ACT COAL")
        plan_sr = total("PLAN SR")
        actual_sr = total("SR")
        productivity = avg("PDTY")

        # 3. Weather Review (Satu baris cuaca)
        weather = {
            "rain_hours": total("RAIN"),
            "slippery_hours": total("SLIPERRY"),
            "foggy_hours": total("FOGGY"),
            "rain_intensity": avg("RAIN INT"),
            "rain_freq": total("RAIN X"),
            "slippery_freq": total("SLIPERRY X"),
            "foggy_freq": total("FOGGY X"),
        }

        # 4. Fleet & Equipment Availability (Pink)
        fleet = {
            "plan_fleet": avg("PLAN FLEET RUNNING "),
            "actual_fleet": avg("ACTUAL FLEET RUNNING "),
            "plan_pa_prod": avg("PLAN PA PROD EQP"),
            "actual_pa_prod": avg("ACT PA PROD EQP "),
            "plan_pa_supp": avg("PLAN PA SUPPORT EQP"),
            "actual_pa_supp": avg("ACT PA SUPPORT EQP "),
        }

        # 5. Utilization (UA & UO)
        utilization = {
            "plan_ua": avg("PLAN UA"),
            "actual_ua": avg("ACT UA"),
            "plan_uo": avg("PLAN UO"),
            "actual_uo": avg("ACT UO"),
        }

        return jsonify(
            range={
                "start": min(row["parsed_date"] for row in rows).date().isoformat(),
                "end": max(row["parsed_date"] for row in rows).date().isoformat(),
            },
            materials=materials,
            production={
                "ob": {"plan": plan_ob, "actual": actual_ob, "progress": round(actual_ob / plan_ob * 100, 1) if plan_ob else 0},
                "coal": {"plan": plan_coal, "actual": actual_coal, "progress": round(actual_coal / plan_coal * 100, 1) if plan_coal else 0},
                "sr": {"plan": plan_sr, "actual": actual_sr, "progress": round(actual_sr / plan_sr * 100, 1) if plan_sr else 0},
                "productivity": productivity,
            },
            weather=weather,
            fleet=fleet,
            utilization=utilization,
        )
    except (OSError, ValueError, requests.RequestException) as error:
        return jsonify(error=f"Data produksi tidak dapat dimuat: {error}"), 502


@app.get("/api/reports")
def list_reports() -> Any:
    """List downloadable Excel reports ordered from newest to oldest."""
    if not LOG_DIR.exists():
        return jsonify(reports=[])
    reports = [
        {
            "filename": path.name,
            "modified_at": datetime.fromtimestamp(path.stat().st_mtime, LOG_TIMEZONE).isoformat(),
        }
        for path in sorted(LOG_DIR.glob("*.xlsx"), key=lambda item: item.stat().st_mtime, reverse=True)
    ]
    return jsonify(reports=reports)


@app.get("/api/reports/<path:filename>")
def report_data(filename: str) -> tuple[Any, int] | Any:
    """Transform one generated workbook into dashboard-ready aggregate data."""
    from openpyxl import load_workbook

    try:
        path = report_file_path(filename)
        workbook = load_workbook(path, read_only=True, data_only=True)
        required_sheets = {"Ringkasan", "Ringkasan Kendaraan", "Kecepatan"}
        if not required_sheets.issubset(workbook.sheetnames):
            raise ValueError("Workbook bukan laporan Pit Dispatch Monitoring yang valid.")

        summary = {
            key: value
            for key, value in workbook["Ringkasan"].iter_rows(min_row=3, values_only=True)
            if key is not None
        }
        vehicles = []
        for row in workbook["Ringkasan Kendaraan"].iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            vehicles.append(
                {
                    "tracker_id": int(row[0]),
                    "first_frame": int(row[1]),
                    "last_frame": int(row[2]),
                    "samples": int(row[3]),
                    "average_confidence": round(float(row[4] or 0), 2),
                    "max_speed": int(row[5] or 0),
                }
            )
        speed_rows = []
        for row in workbook["Kecepatan"].iter_rows(min_row=2, values_only=True):
            if row[0] is not None and row[6] is not None:
                speed_rows.append(
                    {
                        "time_seconds": float(row[1] or 0),
                        "segment": str(row[3] or "Tidak diketahui"),
                        "speed": float(row[6]),
                    }
                )
        speed_values = [item["speed"] for item in speed_rows]
        speed_buckets = [("0-20", 0, 20), ("21-40", 21, 40), ("41-60", 41, 60), ("61-80", 61, 80), ("81+", 81, float("inf"))]
        distribution = [
            {"label": label, "count": sum(low <= speed <= high for speed in speed_values)}
            for label, low, high in speed_buckets
        ]
        timeline: dict[int, list[float]] = defaultdict(list)
        for item in speed_rows:
            timeline[int(item["time_seconds"] // 60) * 60].append(item["speed"])
        return jsonify(
            filename=path.name,
            summary=summary,
            metrics={
                "vehicles": len(vehicles),
                "detections": int(summary.get("Jumlah deteksi", 0)),
                "measurements": len(speed_rows),
                "average_speed": round(sum(speed_values) / len(speed_values), 1) if speed_values else 0,
                "max_speed": round(max(speed_values), 1) if speed_values else 0,
            },
            distribution=distribution,
            timeline=[
                {"minute": minute // 60, "average_speed": round(sum(speeds) / len(speeds), 1)}
                for minute, speeds in sorted(timeline.items())
            ],
            segments=[{"name": name, "count": count} for name, count in Counter(item["segment"] for item in speed_rows).most_common()],
            vehicles=sorted(vehicles, key=lambda item: item["max_speed"], reverse=True)[:100],
        )
    except FileNotFoundError:
        return jsonify(error="Laporan tidak ditemukan."), 404
    except (OSError, ValueError, TypeError, KeyError) as error:
        return jsonify(error=f"Laporan tidak dapat dibaca: {error}"), 400


@app.errorhandler(413)
def file_too_large(_: Any) -> tuple[Any, int]:
    """Return a helpful response when an upload exceeds the size limit."""
    return jsonify(error="Ukuran video melebihi batas 2 GB."), 413


if __name__ == "__main__":
    app.run(host="0.0.0.0", debug=True, port=int(os.environ.get("PORT", "5000")))
